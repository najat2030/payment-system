import streamlit as st
import pandas as pd
import os
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(layout="wide")

DB = "database.xlsx"
PRIMARY = "#8B2C2C"   # نبيتي
GREEN = "#0B6B3A"     # أخضر ملكي

# ================== DB INIT ==================
if not os.path.exists(DB):
    pd.DataFrame(columns=["Account","Spoc","Mobile","Previous","Invoice_April_2026","Type"])\
        .to_excel(DB, sheet_name="master", index=False)

st.title("📊 Telecom Payment System")

# ================== UPLOAD ==================
st.subheader("📂 رفع البيانات الأساسية")
file = st.file_uploader("Upload Excel", type=["xlsx"])

if file:
    df_upload = pd.read_excel(file)
    if "Type" not in df_upload.columns:
        df_upload["Type"] = "Normal"
    df_upload = df_upload.fillna(0)
    df_upload = df_upload.replace("No Show", 0)

    df_upload.to_excel(DB, sheet_name="master", index=False)
    st.success("تم حفظ البيانات بنجاح")
    st.rerun()

master = pd.read_excel(DB, sheet_name="master")

# ================== INPUT ==================
st.subheader("✍️ تحديث القيم الحالية")

input_df = st.data_editor(
    master[master["Type"] != "NonPayment"][["Mobile","Previous"]]
    .rename(columns={"Previous":"Current"}),
    use_container_width=True
)

# ================== UPDATE ==================
if st.button("🚀 تحديث النظام"):

    df = master.merge(input_df, on="Mobile", how="left")
    df["Current"] = df["Current"].fillna(df["Previous"])

    df["Paid"] = df.apply(
        lambda row: 0 if row["Type"] == "NonPayment"
        else row["Previous"] - row["Current"], axis=1
    )

    df["Overpayment"] = df["Current"].apply(lambda x: abs(x) if x < 0 else 0)

    df["Collection"] = df.apply(
        lambda row: 0 if row["Type"] == "NonPayment"
        else min(row["Paid"], row["Previous"]), axis=1
    )

    normal_df = df[df["Type"] != "NonPayment"]

    # ================== DASHBOARD ==================
    c1, c2, c3 = st.columns(3)
    c1.metric("💰 المدفوع", f"{int(normal_df['Paid'].sum()):,}")
    c2.metric("🟢 التحصيل", f"{int(normal_df['Collection'].sum()):,}")
    c3.metric("🔴 Overpayment", f"{int(normal_df['Overpayment'].sum()):,}")

    # ================== FORMAT ==================
    for col in ["Current","Paid","Collection","Overpayment"]:
        df[col] = df[col].apply(lambda x: f"{x:,.2f}" if isinstance(x,(int,float)) else x)

    # ================== HTML TABLE ==================
    def build_html(df):
        html = f"""
        <style>
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 15px;
        }}
        th {{
            background-color: {PRIMARY};
            color: white;
            padding: 10px;
        }}
        td {{
            padding: 8px;
            text-align: center;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        .neg {{
            color: red;
            font-weight: bold;
        }}
        .nonpay {{
            background-color: #d4edda;
        }}
        </style>
        <table>
        <tr>{"".join([f"<th>{c}</th>" for c in df.columns])}</tr>
        """

        for _, row in df.iterrows():
            cls = "nonpay" if row["Type"] == "NonPayment" else ""
            html += f"<tr class='{cls}'>"
            for col in df.columns:
                val = row[col]
                if isinstance(val, str) and val.startswith("-"):
                    html += f"<td class='neg'>{val}</td>"
                else:
                    html += f"<td>{val}</td>"
            html += "</tr>"

        html += "</table>"
        return html

    st.subheader("📋 التقرير الأساسي")
    st.markdown(build_html(df), unsafe_allow_html=True)

    # ================== PAYMENT REPORT ==================
    def format_val(row):
        if row["Type"] == "NonPayment":
            return f"Non Payment - {row['Previous']}"
        return row["Current"]

    df["Display"] = df.apply(format_val, axis=1)

    payment = df[["Display","Mobile","Account"]]
    payment.columns = ["مبلغ مستحق","Phone","Account"]

    payment = payment.sort_values(by="مبلغ مستحق", ascending=False)

    total = normal_df["Current"].sum()

    st.subheader("📄 تقرير الدفع")
    st.markdown(build_html(payment), unsafe_allow_html=True)
    st.metric("💰 إجمالي المطلوب", f"{int(total):,} جنيه")

    # ================== EXCEL EXPORT (Styled) ==================
    wb = Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color="8B2C2C", end_color="8B2C2C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(payment.columns.tolist())

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in payment.values:
        ws.append(list(row))

    wb.save("report.xlsx")

    with open("report.xlsx","rb") as f:
        st.download_button("📥 Excel", f)

    # ================== PDF ==================
    doc = SimpleDocTemplate("report.pdf")
    data = [payment.columns.tolist()] + payment.values.tolist()
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor(PRIMARY)),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('GRID',(0,0),(-1,-1),1,colors.black)
    ]))
    doc.build([table])

    with open("report.pdf","rb") as f:
        st.download_button("📄 PDF", f)

    # ================== IMAGE ==================
    fig, ax = plt.subplots(figsize=(10,6))
    ax.axis('off')
    ax.table(cellText=payment.values, colLabels=payment.columns, loc='center')
    plt.savefig("report.png")

    with open("report.png","rb") as f:
        st.download_button("🖼️ صورة", f)
